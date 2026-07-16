"""Expense-type, expense, sub-expense-type and expense-report routes
(moved verbatim from app.py)."""
import csv
import os
import math
import calendar
from io import StringIO
from datetime import datetime, date

from flask import Blueprint, Response, render_template, request, redirect, url_for, session, flash, jsonify, abort
from pytz import timezone
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from app.config import INVOICE_FOLDER, EXPENSE_TEMPLATE_PATH
from app.extensions import db

expenses_bp = Blueprint('expenses', __name__)


@expenses_bp.route('/admin/expense_types', methods=['GET', 'POST'])
def admin_expense_types():
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    if request.method == 'POST':
        etype = request.form['expense_type'].strip().title()
        if etype:
            db.add_expense_type(etype)
            msg = f'Expense type "{etype}" saved (duplicates are ignored)'
            return redirect(url_for('expenses.admin_expense_types', msg=msg))

    types_ = db.get_expense_types()
    msg = request.args.get('msg')
    return render_template('expenses/admin_expense_types.html', types=types_, msg=msg)

@expenses_bp.route('/admin/delete_expense_type/<int:et_id>')
def delete_expense_type(et_id):
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))
    db.delete_expense_type(et_id)
    flash('Expense type deleted', 'success')
    return redirect(url_for('expenses.admin_expense_types'))

@expenses_bp.route('/admin/edit_expense_type/<int:et_id>', methods=['POST'])
def edit_expense_type(et_id):
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    new_type = request.form['new_expense_type'].strip().title()
    if new_type:
        if db.expense_type_exists(new_type, exclude_id=et_id):
            flash(f'Expense type "{new_type}" already exists.', 'error')
        else:
            db.update_expense_type(et_id, new_type)
            flash(f'Expense type updated to "{new_type}"', 'success')

    return redirect(url_for('expenses.admin_expense_types'))


@expenses_bp.route('/employee/expense', methods=['GET', 'POST'])
def expense():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    if request.method == 'POST':
        expense_date_str = request.form['expense_date']

        # Validate date is in current month
        expense_date = datetime.strptime(expense_date_str, '%Y-%m-%d').date()
        today = date.today()

        if expense_date.month != today.month or expense_date.year != today.year:
            flash('Expense date must be within the current month.', 'error')
            types = db.get_expense_types()
            employees = db.get_employees()
            return render_template('expenses/expense.html', types=types, employees=employees)

        # Handle file upload
        invoice_path = None
        if 'invoice' in request.files:
            file = request.files['invoice']
            if file and file.filename:
                filename = secure_filename(f"{session['user_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                file.save(os.path.join(INVOICE_FOLDER, filename))
                # DB keeps the project-relative path (templates render it as /<path>)
                invoice_path = os.path.join('static/invoices', filename)

        # Get employee_id
        if session['emp_type'] == 'admin' and request.form.get('employee_id'):
            employee_id = int(request.form['employee_id'])
        else:
            employee_id = session['user_id']

        # Get expense_by value
        expense_by = request.form.get('expense_by', '')
        if expense_by == 'Self':
            # Get employee name
            emp = db.get_employee(employee_id)
            expense_by = f"Self ({emp[1]} {emp[2]})"  # Self (FirstName LastName)

        data = {
            'expense_type_id': int(request.form['expense_type_id']),
            'sub_expense_type_id': int(request.form['sub_expense_type_id']) if request.form.get('sub_expense_type_id') else None,
            'employee_id': employee_id,
            'exp_description': request.form['description'][:500],
            'amount': float(request.form['amount']),
            'expense_date': expense_date_str,
            'invoice_path': invoice_path,
            'po_no': request.form.get('po_no', ''),
            'bill_status': request.form.get('bill_status', ''),
            'expense_by': expense_by,
            'manager_id': None,
            'given_by_id': None
        }

        db.add_expense(data)
        flash('Expense request submitted successfully!', 'success')
        return redirect(url_for('expenses.existing_expenses'))

    types = db.get_expense_types()
    employees = db.get_employees() if session['emp_type'] == 'admin' else []

    # Get current month range for date validation
    today = date.today()
    first_day = today.replace(day=1)
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    return render_template('expenses/expense.html',
                         types=types,
                         employees=employees,
                         min_date=first_day.isoformat(),
                         max_date=last_day.isoformat())

@expenses_bp.route('/existing_expenses')
def existing_expenses():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    page = int(request.args.get('page', 1))
    per_page = 15
    offset = (page - 1) * per_page

    filters = []
    params = []

    employee_id = request.args.get('employee_id', '').strip()
    expense_type = request.args.get('expense_type', '').strip()
    status = request.args.get('status', '').strip()
    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()

    ist = timezone('Asia/Kolkata')  # IST timezone

    if session['emp_type'] == 'admin':
        if employee_id:
            filters.append("ex.employee_id = ?")
            params.append(employee_id)
        if expense_type:
            filters.append("et.expense_type = ?")
            params.append(expense_type)
        if status:
            filters.append("ex.status = ?")
            params.append(status)
        if from_date:
            filters.append("DATE(ex.inserted_date) >= ?")
            params.append(from_date)
        if to_date:
            filters.append("DATE(ex.inserted_date) <= ?")
            params.append(to_date)

        where_clause = 'WHERE ' + ' AND '.join(filters) if filters else ''
        total = db.count_expenses(where_clause, tuple(params))
        expenses = db.get_expenses_paginated(where_clause, tuple(params), per_page, offset)
        expense_types = db.get_expense_types()
        employees = db.get_employees(status_filter='active')

    else:
        base = "WHERE ex.employee_id = ?"
        params = [session['user_id']]
        if status:
            base += " AND ex.status = ?"
            params.append(status)
        if from_date:
            base += " AND DATE(ex.inserted_date) >= ?"
            params.append(from_date)
        if to_date:
            base += " AND DATE(ex.inserted_date) <= ?"
            params.append(to_date)

        total = db.count_expenses(base, tuple(params))
        expenses = db.get_expenses_paginated(base, tuple(params), per_page, offset)
        expense_types = []
        employees = []

    # Indices from your table structure
    IDX_INSERTED_DATE = 7
    IDX_APPROVED_DATE = 10

    # Convert tuple → list and adjust date/time
    converted_expenses = []
    for e in expenses:
        e = list(e)  # Make it mutable
        try:
            if e[IDX_INSERTED_DATE]:
                utc_time = datetime.fromisoformat(e[IDX_INSERTED_DATE])
                e[IDX_INSERTED_DATE] = utc_time.astimezone(ist).strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            pass

        try:
            if e[IDX_APPROVED_DATE]:
                utc_time = datetime.fromisoformat(e[IDX_APPROVED_DATE])
                e[IDX_APPROVED_DATE] = utc_time.astimezone(ist).strftime('%Y-%m-%d %H:%M:%S')
            else:
                e[IDX_APPROVED_DATE] = 'Not set'
        except Exception:
            e[IDX_APPROVED_DATE] = 'Not set'

        converted_expenses.append(e)

    total_pages = math.ceil(total / per_page)

    return render_template('expenses/existing_expenses.html',
                           expenses=converted_expenses,
                           expense_types=expense_types,
                           employees=employees,
                           page=page,
                           total_pages=total_pages)

@expenses_bp.route('/export_expenses')
def export_expenses():
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    expenses = db.get_expenses()  # Use all data, no filter

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow(['SlNo', 'ExpType', 'ExpnDate', 'Amt', 'Name', 'ReqDate', 'Status', 'ApprovedBy', 'Comments'])

    for i, ex in enumerate(expenses, start=1):
        writer.writerow([
            i,
            ex[1],                                # ExpType
            ex[11] or '',                         # ExpnDate
            ex[8],                                # Amount
            ex[2],                                # Name
            ex[7],                                # ReqDate
            ex[4],                                # Status
            ex[12] if len(ex) > 12 else '',       # ApprovedBy (added to SELECT)
            ex[5] or ''                           # Comments
        ])

    # Send response
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            "Content-Disposition": "attachment; filename=expenses_export.csv"
        }
    )


@expenses_bp.route('/api/expense/<int:exp_id>')
def get_expense_detail(exp_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    exp = db.get_expense_by_id(exp_id)
    if not exp:
        return jsonify({'error': 'Not found'}), 404

    ist = timezone('Asia/Kolkata')

    # Handle requested_date
    try:
        if exp[8]:
            req_dt = datetime.fromisoformat(exp[8]).astimezone(ist).strftime('%Y-%m-%d %H:%M:%S')
        else:
            req_dt = 'Not set'
    except Exception:
        req_dt = exp[8]

    # Handle approved_date
    try:
        if exp[11]:
            app_dt = datetime.fromisoformat(exp[11]).astimezone(ist).strftime('%Y-%m-%d %H:%M:%S')
        else:
            app_dt = 'Not set'
    except Exception:
        app_dt = 'Not set'

    return jsonify({
        'expense_id': exp[0],
        'type': exp[1],
        'sub_type': exp[2] if exp[2] else 'N/A',
        'emp_name': exp[3],
        'description': exp[4],
        'status': exp[5],
        'approver_comments': exp[6],
        'final_comments': exp[7],
        'requested_date': req_dt,
        'amount': exp[9],
        'employee_id': exp[10],
        'approved_date': app_dt,
        'po_no': exp[12] if exp[12] else 'N/A',
        'bill_status': exp[13] if exp[13] else 'N/A',
        'expense_by': exp[14] if exp[14] else 'N/A'
    })

@expenses_bp.route('/admin/expense/<int:exp_id>/<action>', methods=['POST'])
def expense_approve(exp_id, action):
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    if action not in ('approved', 'rejected'):
        abort(400)

    comments = request.form.get('approver_comments', '')[:200]
    approved_by = request.form.get('approved_by', '').strip() or None
    db.update_expense_status(exp_id, action, comments, session['user_id'], approved_by)
    flash(f'Expense {action}', 'success')
    return redirect(url_for('expenses.existing_expenses'))

@expenses_bp.route('/admin/expense/generate_report/<int:emp_id>')
def generate_expense_report(emp_id):
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    # Get employee details
    employee = db.get_employee(emp_id)
    if not employee:
        flash('Employee not found', 'error')
        return redirect(url_for('expenses.existing_expenses'))

    emp_name = f"{employee[1]} {employee[2]}"  # first_name last_name

    # Get current month expenses for this employee
    today = date.today()
    first_day = today.replace(day=1)
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    # Get expenses
    conn = db.get_connection()
    cursor = conn.cursor()
    expenses = cursor.execute('''
        SELECT ex.expense_id, et.expense_type, st.sub_expense_type,
               e.first_name||' '||e.last_name AS emp_name,
               ex.exp_description, ex.status,
               ex.inserted_date, ex.amount,
               ex.expense_date,
               ex.po_no,
               ex.bill_status,
               ex.expense_by
        FROM tbl_expenses ex
        JOIN tbl_expense_type et ON et.expense_type_id = ex.expense_type_id
        LEFT JOIN tbl_sub_expense_type st ON st.sub_expense_type_id = ex.sub_expense_type_id
        JOIN tbl_employee e ON e.emp_id = ex.employee_id
        WHERE ex.employee_id = ? AND ex.expense_date BETWEEN ? AND ?
        ORDER BY ex.expense_date
    ''', (emp_id, first_day.isoformat(), last_day.isoformat())).fetchall()
    conn.close()

    # Load template
    template_path = EXPENSE_TEMPLATE_PATH
    wb = load_workbook(template_path)
    ws = wb.active

    # Update employee name and period
    ws['C3'] = emp_name
    ws['C5'] = f"{first_day.strftime('%d-%m-%Y')} to {last_day.strftime('%d-%m-%Y')}"

    # Clear existing data (from row 10 onwards)
    ws.delete_rows(10, ws.max_row - 9)

    # Add expense data starting from row 10
    start_row = 10
    for idx, exp in enumerate(expenses, 1):
        row = start_row + idx - 1

        # Parse expense_date
        try:
            if exp[8]:
                exp_date = datetime.fromisoformat(exp[8]).date()
                formatted_date = exp_date.strftime("%d-%m-%Y")
            else:
                exp_date = None
        except:
            exp_date = None

        ws[f'A{row}'] = idx
        ws[f'B{row}'] = formatted_date
        ws[f'C{row}'] = exp[9] if exp[9] else ''  # po_no
        ws[f'D{row}'] = exp[1]  # expense_type
        ws[f'E{row}'] = exp[2] if exp[2] else ''  # sub_expense_type
        ws[f'F{row}'] = exp[4] if exp[4] else ''  # description
        ws[f'G{row}'] = exp[10] if exp[10] else ''  # bill_status
        ws[f'H{row}'] = exp[7]  # amount
        ws[f'I{row}'] = exp[11] if exp[11] else ''  # expense_by

    # Save to BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create filename
    filename = f"Expense_Report_{emp_name.replace(' ', '_')}_{today.strftime('%B_%Y')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

@expenses_bp.route('/admin/expense/report')
def generate_expense_report_form():
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    emp_id = request.args.get('emp_id')
    if emp_id:
        return redirect(url_for('expenses.generate_expense_report', emp_id=int(emp_id)))

    flash('Please select an employee', 'error')
    return redirect(url_for('expenses.existing_expenses'))

# ------ Sub Expense types ----------------------------------
@expenses_bp.route('/admin/sub_expense_types', methods=['GET', 'POST'])
def admin_sub_expense_types():
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))

    popup_message = None
    popup_type = None

    if request.method == 'POST':
        expense_type_id = request.form['expense_type_id']
        sub_expense_type = request.form['sub_expense_type'].strip()
        sub_et_id = request.form.get('sub_expense_type_id')

        if not sub_expense_type:
            popup_message = 'Sub-expense type cannot be empty.'
            popup_type = 'error'
        elif sub_et_id:  # Edit
            success = db.update_sub_expense_type(int(sub_et_id), sub_expense_type)
            popup_message = 'Sub-expense type updated successfully.' if success else 'Update failed.'
            popup_type = 'success' if success else 'error'
        else:  # Add
            success, message = db.add_sub_expense_type(int(expense_type_id), sub_expense_type)
            popup_message = message
            popup_type = 'success' if success else 'error'

    edit_type = None
    edit_id = request.args.get('edit_id')
    if edit_id:
        sub_types_all = db.get_sub_expense_types()
        edit_type = next((st for st in sub_types_all if str(st[0]) == edit_id), None)

    sub_types_all = db.get_sub_expense_types()
    expense_types = db.get_expense_types()

    return render_template('expenses/admin_sub_expense_types.html',
                         sub_expense_types=sub_types_all,
                         expense_types=expense_types,
                         popup_message=popup_message,
                         popup_type=popup_type,
                         edit_type=edit_type)

@expenses_bp.route('/admin/delete_sub_expense_type/<int:sub_et_id>')
def delete_sub_expense_type(sub_et_id):
    if 'user_id' not in session or session['emp_type'] != 'admin':
        return redirect(url_for('auth.login'))
    db.delete_sub_expense_type(sub_et_id)
    flash('Sub-expense type deleted', 'success')
    return redirect(url_for('expenses.admin_sub_expense_types'))

@expenses_bp.route('/api/sub_expense_types/<int:expense_type_id>')
def get_sub_expense_types_api(expense_type_id):
    """API endpoint to get sub-expense types for a given expense type"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    sub_types = db.get_sub_expense_types(expense_type_id)
    result = [{'id': st[0], 'name': st[1]} for st in sub_types]
    return jsonify(result)


@expenses_bp.route('/travel_policy')
def travel_policy():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    return render_template('expenses/travel_policy.html')

